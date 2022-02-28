import os
from openpyxl import Workbook
from module import calculate_score

k = 0
os.chdir("integration")
wb = Workbook()
ws = wb.active
ws.title = "점수 통계"
imglist = os.listdir()
imglist.remove('pigmentation_score.xlsx')
imglist.remove('pigmentation_score2.xlsx')
start_cnt = 2
for i in range(1, 101):
    cnt = 0
    k += 1
    ws.cell(row=i, column=1, value=i)
    for j in range(len(imglist)):
        if int(sorted(imglist)[j][0:3]) == i:
            ws.cell(row=i, column=start_cnt + cnt, value=calculate_score(sorted(imglist)[j]))
            cnt += 1

    print(k)
wb.save("pigmentation_score.xlsx")
